from selenium import webdriver
import os
import time
from selenium.webdriver.chrome.options import Options
import openpyxl


def getpath():
    #获取当前绝对路径
    dirpath = os.path.dirname(os.path.abspath(__file__))
    fileName = 'boosRecruit' + '.xlsx'
    filePath = os.path.join(dirpath,fileName)
    return  filePath

def getSheet():
    path = getpath()
    if os.path.exists(path) is True:
        print('存在文件,获得excle对象')
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.cell(1, 9, value='职责描述')
        sheet.cell(1,10,value='任职要求')
        return workbook,sheet
    else:
        print('不存在文件请运行BOSS爬虫')

def getDriver(url):
    # 通过cmd启动一个自定义浏览器
    chrome_options = Options()
    options = 'chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfil'
    chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    driverPath = 'F:\python_test\\test_yy_cong\\drivers\\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driverPath,chrome_options=chrome_options)
    driver.get(url)
    return driver


def sayhellow():
    path  = getpath()
    sheet = getSheet()
    workbook = sheet[0]
    sheet = sheet[1]
    for i in range(1,sheet.max_row):
        url = sheet.cell(i+1,6).value
        driver = getDriver(url)
        driver.implicitly_wait(30)
        if sheet.cell(i+1,9).value is None:
            jobsec = driver.find_element_by_xpath('//*[@class="job-sec"]//*[@class="text"]').text
            jobaddress =driver.find_element_by_xpath('//*[@class="job-location"]//*[@class="location-address"]').text
            jobsec = jobsec.replace('任职要求','**任职要求').split('**')
            sheet.cell(i + 1,9,value=jobsec[0])
            sheet.cell(i + 1,10, value=jobsec[1])
            sheet.cell(i + 1,4, value=jobaddress)
            workbook.save(path)
            if sheet.cell(i + 1,11).value != '是':
                print(sheet.cell(i + 1,11).value)
                driver.find_element_by_xpath('//*[@id="main"]/div[1]/div/div/div[3]/div[1]/a').click()
                time.sleep(2)
                driver.find_element_by_xpath('//*[@class="dialog-footer"]//*[@ka=\'dialog_confirm\']').click()
                time.sleep(2)
                text=''
                driver.find_element_by_xpath('//*[@class="chat-im chat-editor"]//*[@class=\'chat-input\']').send_keys(text)
                time.sleep(1)
                send = driver.find_element_by_xpath('//*[@class="chat-im chat-editor"]//*[@class=\'btn btn-primary btn-send\']')
                print(send.is_enabled())
                if send.is_enabled() is True:
                    send.click()
                    sheet.cell(1,11,value='是否已打招呼')
                    sheet.cell(i + 1,11,value='是')
                    workbook.save(path)
                    driver.close()
            else:
                sheet.cell(1, 11, value='是否已打招呼')
                sheet.cell(i + 1, 11, value='是')
                workbook.save(path)
                driver.close()





if __name__ == '__main__':




