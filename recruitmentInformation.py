from selenium import webdriver
import os
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import openpyxl


# 获取当前时间
def getTieme():
    now = int(time.time())
    timeArray = time.localtime(now)
    currentTime = time.strftime('%Y-%m-%d %H:%M:%S', timeArray)
    return  currentTime

def getpath():
    #获取当前绝对路径
    dirpath = os.path.dirname(os.path.abspath(__file__))
    fileName = 'boosRecruit' + '.xlsx'
    filePath = os.path.join(dirpath,fileName)
    return  filePath


#判断文件是否存在，不存在则创建并返回表对象和sheet对象
def getSheet():
    path = getpath()
    if os.path.exists(path) is True:
        print('存在文件,获得excle对象')
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        return workbook,sheet
    else:
        print('不存在文件,创建文件并获得excle对象')
        wb = openpyxl.Workbook()
        wb.save(path)
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.cell(1, 1, value='公司名称')
        sheet.cell(1, 2, value='薪资')
        sheet.cell(1, 3, value='联系人')
        sheet.cell(1, 4, value='地点')
        sheet.cell(1, 5, value='发布时间')
        sheet.cell(1, 6, value='具体url')
        sheet.cell(1, 7, value='公司介绍网址')
        sheet.cell(1, 8, value='写入时间')
        return workbook,sheet


#获取初使话浏览器对象
def getDriver(url):
    # 通过cmd启动一个自定义浏览器
    chrome_options = Options()
    options = 'chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfil'
    chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    os.popen(options)
    driverPath = 'F:\python_test\\test_yy_cong\\drivers\\chromedriver.exe'
    driver = webdriver.Chrome(executable_path=driverPath,chrome_options=chrome_options)
    driver.get(url)
    return driver


#获取数据
def getData(url,jobname):
    sheet = getSheet()
    workbook = sheet[0]
    sheet = sheet[1]
    driver = getDriver(url)
    driver.implicitly_wait(30)
    driver.maximize_window()
    action = ActionChains(driver)
    path = getpath()
    # 输入框
    jobnameinput = driver.find_element_by_xpath('//*[@class="ipt-wrap"]//*[@name=\'query\']')
    jobnameinput.clear()
    jobnameinput.send_keys(jobname)
    time.sleep(0.5)
    jobnameinput.send_keys(Keys.ENTER)
    time.sleep(3)
    releaseTimeIcon = driver.find_element_by_xpath('//*[@class="filter-select-box"]//div[6]//i')
    action.move_to_element(releaseTimeIcon).perform()
    time.sleep(0.5)
    # 通过时间进行筛选 不限，一天,三天，七天，15天，一个月以内/可根据自己需要修改定位直接修改li内索引元素就行了从1开始的
    driver.find_element_by_xpath('//*[@class="filter-select-box"]//div[6]//li[6]').click()
    for g in range(1,1000):
        corporateName  = driver.find_elements_by_xpath('//*[@class="job-list"]//*[@class="company-text"]//a')
        contacts  = driver.find_elements_by_xpath('//*[@class="job-list"]//*[@class="info-publis"]/h3')
        address = driver.find_elements_by_xpath('//*[@class="job-list"]//*[@class="info-primary"]//p')
        releaseTime= driver.find_elements_by_xpath('//*[@class="job-list"]//*[@class="info-publis"]//p')
        salary = driver.find_elements_by_xpath('//*[@class="job-list"]//*[@class="info-primary"]//h3/a')
        for i in range(0,len(corporateName)):
            corporateNames = corporateName[i].text
            companyIntroductionUrl  = corporateName[i].get_attribute('href')
            contactss = contacts[i].text
            addresss = address[i].text
            releaseTimes = releaseTime[i].text
            salarys = salary[i].text
            detailsUrl  = salary[i].get_attribute('href')
            writetimes = getTieme()
            rows = sheet.max_row + 1
            sheet.cell(rows, 1,value=corporateNames)
            sheet.cell(rows, 2, value=salarys)
            sheet.cell(rows, 3, value=contactss)
            sheet.cell(rows, 4, value=addresss)
            sheet.cell(rows, 5, value=releaseTimes)
            sheet.cell(rows, 6, value=detailsUrl)
            sheet.cell(rows, 7, value=companyIntroductionUrl)
            sheet.cell(rows, 8, value=writetimes)
            workbook.save(path)
            if i >= len(corporateName) - 1 :
                print('爬取第'+str(g) +'页完成，进行下一页')
                break
        try:
            next = driver.find_element_by_xpath('//*[@class="job-list"]//*[@ka="page-next"]')
            if next.get_attribute('class') == 'next':
                print('点击下一页')
                next.click()
            else:
                driver.close()
                print('一共获得' + str(g) + '页数据')
                break
        except Exception as e:
            driver.close()
            print('没有获得翻页按钮')
            raise e






if __name__ == '__main__':
    url = 'https://www.zhipin.com/'
    jobname = '软件测试工程师'
    getData(url,jobname)

